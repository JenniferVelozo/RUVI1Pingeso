import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE","backendRUVI1.settings")
import pandas as pd
from numpy import arange
import django
django.setup()
from gestionPacientes.models import Resumen
from openpyxl.utils.cell import get_column_interval
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import json

# va a la BD a buscarlo
def resumen_search_to_excel():
    
    resumenQuery=Resumen.objects.all()
    resumen=pd.DataFrame()
    cama=[]
    rut=[]
    nombrePaciente=[] 
    estancia=[] 
    criterio=[] 
    diagnostico1=[] 
    diagnostico2 =[]
    ir_grd=[] 
    emNorma=[] 
    pcSuperior=[] 
    pesoGRD=[] 
    nombreServicio=[] 
    servicio=[] 
    flag_diag =[]
    for paciente in resumenQuery:
        cama.append(paciente.cama) 
        rut.append(paciente.rut) 
        nombrePaciente.append(paciente.nombrePaciente) 
        estancia.append(paciente.estancia) 
        criterio.append(paciente.criterio) 
        diagnostico1.append(paciente.diagnostico1) 
        diagnostico2.append(paciente.diagnostico2)
        ir_grd.append(paciente.ir_grd) 
        emNorma.append(paciente.emNorma) 
        pcSuperior.append(paciente.pcSuperior) 
        pesoGRD.append(paciente.pesoGRD) 
        nombreServicio.append(paciente.nombreServicio) 
        servicio.append(paciente.servicio.id) 
        flag_diag.append(paciente.flag_diag)

    resumen=resumen.assign(rut=rut, nombrePaciente=nombrePaciente, cama=cama, estancia=estancia,
    diagnostico1=diagnostico1, diagnostico2=diagnostico2, ir_grd=ir_grd, emNorma=emNorma, 
    pcSuperior=pcSuperior, pesoGRD=pesoGRD, nombreServicio=nombreServicio, servicio=servicio, criterio=criterio, flag_diag=flag_diag)
    
    print(resumen)
    nombreArchivo='Gestion de Pacientes.xlsx'
    nombreArchivoR='\Gestion de Pacientes.xlsx'
    resumen.to_excel(nombreArchivo, sheet_name='Resumen de pacientes')
    estiloExcel(nombreArchivo)
    return os.path.dirname(os.path.abspath(__file__)) + nombreArchivoR
    

# establece el tamaño de las celdas y los colores para que no se sobrepongan
def estiloExcel(nombre):
    informe = openpyxl.load_workbook(nombre)

    sheet = informe.active 

    def set_width_to(sheet, start, stop, width):
        for col in get_column_interval(start, stop):
            sheet.column_dimensions[col].width = width

    set_width_to(sheet, "A", "A", width=5)
    set_width_to(sheet, "B", "B", width=14)
    set_width_to(sheet, "C", "C", width=52)
    set_width_to(sheet, "D", "D", width=10)
    set_width_to(sheet, "E", "E", width=9)
    set_width_to(sheet, "F", "G", width=84)
    set_width_to(sheet, "H", "K", width=10)
    set_width_to(sheet, "L", "L", width=40)
    set_width_to(sheet, "M", "O", width=9)

    encabezados=["A1","B1","C1","D1","E1","F1","G1","H1","I1","J1","K1","L1","M1","N1","O1"]
    for encabezado in encabezados:
        celda = sheet[encabezado]
        celda.fill =  PatternFill("solid", fgColor="D9D9D9")
    informe.save(nombre)

#lo pide en formato json como entrada.
def resumen_to_excel(resumenJSON):
    cama=[]
    rut=[]
    nombrePaciente=[] 
    estancia=[] 
    criterio=[] 
    diagnostico1=[] 
    diagnostico2 =[]
    ir_grd=[] 
    emNorma=[] 
    pcSuperior=[] 
    pesoGRD=[] 
    nombreServicio=[] 
    servicio=[] 
    flag_diag =[]
    for paciente in resumenJSON:
        cama.append(paciente['cama']) 
        rut.append(paciente['rut']) 
        nombrePaciente.append(paciente['nombrePaciente']) 
        estancia.append(paciente['estancia']) 
        criterio.append(paciente['criterio']) 
        diagnostico1.append(paciente['diagnostico1']) 
        diagnostico2.append(paciente['diagnostico2'])
        ir_grd.append(paciente['ir_grd']) 
        emNorma.append(paciente['emNorma']) 
        pcSuperior.append(paciente['pcSuperior']) 
        pesoGRD.append(paciente['pesoGRD']) 
        nombreServicio.append(paciente['nombreServicio']) 
        servicio.append(paciente['servicio']) 
        flag_diag.append(paciente['flag_diag'])

    resumen= pd.DataFrame()
    resumen=resumen.assign(rut=rut, nombrePaciente=nombrePaciente, cama=cama, estancia=estancia,
    diagnostico1=diagnostico1, diagnostico2=diagnostico2, ir_grd=ir_grd, emNorma=emNorma, 
    pcSuperior=pcSuperior, pesoGRD=pesoGRD, nombreServicio=nombreServicio, servicio=servicio, criterio=criterio, flag_diag=flag_diag)
    

    nombreArchivo='Gestion de Pacientes.xlsx'
    nombreArchivoR='\Gestion de Pacientes.xlsx'
    print(resumen)
    resumen.to_excel(nombreArchivo, sheet_name='Resumen de pacientes')
    estiloExcel(nombreArchivo)
    return os.path.dirname(os.path.abspath(__file__)) + nombreArchivoR

if __name__=='__main__':
    #ruta= resumen_to_excel()
    #print(ruta)

    #dummy de prueba
    resumenQuery=Resumen.objects.all()
    resumen=[]
    for paciente in resumenQuery:
        jsonAux={
        "cama": paciente.cama,
        "rut": paciente.rut,
        "nombrePaciente": paciente.nombrePaciente,
        "estancia": paciente.estancia,
        "criterio": paciente.criterio,
        "diagnostico1": paciente.diagnostico1,
        "diagnostico2": paciente.diagnostico2,
        "ir_grd": paciente.ir_grd,
        "emNorma": paciente.emNorma,
        "pcSuperior": paciente.pcSuperior,
        "pesoGRD": paciente.pesoGRD,
        "nombreServicio": paciente.nombreServicio,
        "servicio": paciente.servicio.id,
        "flag_diag": paciente.flag_diag
        }
        resumen.append(jsonAux)
    
    for j in resumen:
        print(j)
        print(",")

    #se entrega json con el resumen.
    ruta= resumen_to_excel(resumen)
    print(ruta)