
from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.decorators import api_view
from django.core import serializers
from gestionPacientes.models import *
from gestionPacientes.df import *
from .serializers import *
from django.http import HttpResponse, JsonResponse
from openpyxl.utils.cell import get_column_interval
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import json
from rest_framework import filters
from rest_framework import status, mixins, generics, viewsets

from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from gestionPacientes.serializers import SendPasswordResetEmailSerializer, UserChangePasswordSerializer, UserLoginSerializer, UserPasswordResetSerializer, UserProfileSerializer, UserRegistrationSerializer
from django.contrib.auth import authenticate
from gestionPacientes.renderers import UserRenderer
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import IsAuthenticated
import os
from django import forms
from gestionPacientes.loadCSV import *
import pandas as pd


class UploadFileForm(forms.Form):
    file = forms.FileField()


@api_view(['POST'])
def subir(request,carga):
    '''Funcion para la carga de archivos de pacientes, cie10 y prestaciones'''
    #respuesta de retorno
    respuesta=[{"Response": "Archivo cargado correctamente. 200_OK", "cargado": True}]
    respuesta2=[{"Response": "Archivo no cargado. 400 BAD REQUEST", "cargado": False}]
    respuesta3=[{"ErrorFormato": "Archivo no cargado. El archivo no cumple el formato. BAD REQUEST", "cargado": False}]
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        #si el formato es valido
        if form.is_valid():
            #se carga el archivo
            valor,retorno=handle_uploaded_file(request.FILES['file'],carga)
            if valor:
                return JsonResponse(respuesta, safe=False, status=status.HTTP_200_OK)
            else:
                if len(retorno)==0:
                    return JsonResponse(respuesta3, safe=False)
                else:
                    #si el archivo tiene diagnosticos no registrados avisa.
                    respuesta4=[{"ErrorDiagnostico": "Archivo no cargado. El archivo tiene un diagnostico que no se encuentra en la base de datos.\n"+
                    "Paciente: "+retorno[0]+" "+retorno[2] +
                    "\nDiagnostico: " + retorno[1]+
                    "\nBAD REQUEST", "cargado": False}]
                    print(respuesta4)
                    return JsonResponse(respuesta4, safe=False)
    else:
        form = UploadFileForm()
    return JsonResponse(respuesta2, safe=False)

  

def handle_uploaded_file(f,carga): 
    '''Dependiendo del archivo de entrada se cargan los datos correspondientes.'''
    path = os.path.dirname(os.path.realpath(__file__))

    if carga=="pacientes":
        path=path+'\\PACIENTES.csv'
    if carga=="CIE10GRD":
        path=path+'\\CIE10-GRD.xlsm'
    if carga=='pendientes':
        path=path+'\\PRESTACIONES_CAUSAS.xlsx'
    
    #se sobreescribe el archivo base de la app con el archivo cargado.
    file = open(path, "w")
    file.close()
    with open(path, 'wb+') as destination:  
      for chunk in f.chunks():  
          destination.write(chunk)  
    #dependiendo del archivo cargado se ejecutan las funciones de los archivos df.py y loadCSV.py
    if carga=="pacientes":
        print("pacientes")
        return leerDf()
    if carga=="CIE10GRD" and f.name=='CIE10-GRD.xlsm':
        print("cie10-norma")
        load_CIE10_GRD(path)
        return True,[]
    if carga=='pendientes' and f.name=='PRESTACIONES_CAUSAS.xlsx':
        print("pendientes")
        load_prestaciones(path)
        return True, []
    return False, []


    
@api_view(['POST'])
def setPendientes(request):
    '''Funcion para agregar pendientes a un paciente.'''
    data=request.data
    idPendientes=data["pendientes"]
    idRes=data["id"]
    pJson=[]
    #para cada id de pendiente entregado
    for idP in idPendientes:
        r=Resumen.objects.get(id=idRes)
        p=Pendientes.objects.get(id=idP)
        flag=True
        #se busca su historico
        try:
            h = Historico.objects.get(fecha=r.updated_at, rut=r.rut, nombrePaciente=r.nombrePaciente,
            diagnostico1Cod=r.diagnostico1Cod, diagnostico2Cod=r.diagnostico2Cod,ir_grd=r.ir_grd,pesoGRD=r.pesoGRD,
            nombreServicio=r.nombreServicio,cama=r.cama, estancia=r.estancia)
        except Historico.DoesNotExist:
            h=None            
            print("no encuentra historico")
            flag=False
        #se guarda el pendiente.
        pJson.append({'id': idP, 'nombre': p.nombrePendiente, 'causa':p.causa })
        r.pendientes.add(p)
        if flag:
            h.pendientes.add(p)
    #se actualizan la fecha de actualizacion y la flag de pendientes.
    fecha = datetime.now()
    r.flag_pend=True
    r.updated_at=fecha
    r.pendientesJson=pJson
    if flag:
        h.flag_pend = True
        h.pendientesJson=pJson
        h.save()
    r.save()
    return HttpResponse(data, content_type='application/json')


def HistoricotoDictionary(historico):
    '''Transforma un reporte historico a dict.'''
    #se crea un json auxiliar donde guardar los datos del historico
    aux={}
    if historico == None:
        return None
    if historico.servicio == None:
        aux["servicio_id"]=None
    else:
        aux["servicio_id"]=historico.servicio.id
    aux["id"]=historico.id
    aux["rut"]= historico.rut
    aux["nombrePaciente"]= historico.nombrePaciente
    aux["fecha"]=historico.fecha
    aux["nombreServicio"]=historico.nombreServicio
    aux["cama"] = historico.cama
    aux["estancia"] = int(historico.estancia) 
    aux["criterio"]=float(historico.criterio)
    aux["diagnostico1"]= historico.diagnostico1
    aux["diagnostico1Cod"]=historico.diagnostico1Cod
    aux["diagnostico2"]= historico.diagnostico2
    aux["diagnostico2Cod"]=historico.diagnostico2Cod
    aux["ir_grd"] = historico.ir_grd
    aux["emNorma"]= float(historico.emNorma)
    aux["pcSuperior"]= int(historico.pcSuperior)
    aux["pesoGRD"] = float(historico.pesoGRD)
    aux["flag_diag"]=historico.flag_diag
    aux["flag_pend"]=historico.flag_pend
    aux["pendientesJson"]=historico.pendientesJson
    return aux


@api_view(['GET'])
def filtrarServicioPendiente(request, fecha, nombreServicio, nombrePendiente):
    '''Get de reportes historicos filtrados fecha, por servicio y por pendientes '''
    historico = Historico.objects.all()
    historico = list(historico)
    # Primero filtramos por fecha
    porFecha = []
    for e in historico:
        if str(e.fecha) == fecha:
            porFecha.append(e)
    #filtro por servicio
    porServicio = []
    for e in porFecha:
        if nombreServicio=="Unidad de gestion de pacientes":
                porServicio.append(e)
        else:
            if e.servicio!=None:
                if e.servicio.nombre == nombreServicio:
                    porServicio.append(e)
    #Filtro por pendiente
    porPendiente = []
    for e in porServicio:        
        if nombrePendiente=="todos":
            porPendiente.append(e)
        else:
            aux=json.dumps({})            
            if e.pendientesJson!=aux:
                for i in e.pendientesJson:
                    if i['nombre'] == nombrePendiente:
                        porPendiente.append(e)
    #resumen final
    listaFinal = []
    for resumen in porPendiente:
        listaFinal.append(HistoricotoDictionary(resumen))
    
    if nombrePendiente == "":
        listaPorServicio = []
        for resumen in porServicio:
            listaPorServicio.append(HistoricotoDictionary(resumen))
            return JsonResponse(listaPorServicio, safe=False, json_dumps_params={'ensure_ascii':False})
    return JsonResponse(listaFinal, safe=False, json_dumps_params={'ensure_ascii':False})


@api_view(['GET'])
def filtrarServicio(request, fecha, nombreServicio):
    '''Get de historicos por fecha y servicio.'''
    historico = Historico.objects.all()
    historico = list(historico)
    # Primero filtramos por fecha
    porFecha = []
    for e in historico:
        if str(e.fecha) == fecha:
            porFecha.append(e)
    #filtro por servicio
    porServicio = []
    for e in porFecha:
        if nombreServicio=="Unidad de gestion de pacientes":
                porServicio.append(e)
        else:
            if e.servicio!=None:
                if e.servicio.nombre == nombreServicio:
                    porServicio.append(e)
    #final.
    listaFinal = []
    for resumen in porServicio:
        listaFinal.append(HistoricotoDictionary(resumen))
    return JsonResponse(listaFinal, safe=False, json_dumps_params={'ensure_ascii':False})

@api_view(['GET'])
def filtrarPendientesPorPaciente(request, id_paciente):
    '''Get del resumen de pacientes filtrado por pendientes.'''
    resumen = Resumen.objects.get(id=int(id_paciente))
    return JsonResponse(resumen.pendientesJson, safe=False, json_dumps_params={'ensure_ascii':False})


@api_view(['GET'])
def reporteMensual(request, year, mes):
    '''Get de reporte mensual por fecha.'''
    
    mensual=ReporteMensual.objects.filter(fecha__year=year, fecha__month=mes)
    #se crea una lista auxiliar con json de cada servicio.
    lista=[]
    for m in mensual:
        mJson={}
        mJson["id"]=m.id
        mJson["fecha"]=m.fecha
        mJson["servicioNombre"]=m.servicioNombre
        mJson["servicioId"]=m.servicio.id
        mJson["em"]=m.em
        mJson["emaf"]=m.emaf
        mJson["iema"]=m.iema
        mJson["peso"]=m.peso
        mJson["iemaInliersMenor"]=m.iemainliersMenor
        mJson["iemaInliersMayor"]=m.iemainliersMayor
        mJson["outliers"]=m.outliers
        mJson["pInt"]=m.pInt
        mJson["pExt"]=m.pExt
        mJson["condP"]=m.condP
        lista.append(mJson)
    return JsonResponse(lista, safe=False,  json_dumps_params={'ensure_ascii':False})

@api_view(['POST'])
def deleteUser(request):
    '''Funcion para eliminar usuario.'''
    data=request.data
    idUser = data["id"]
    User.objects.filter(id=idUser).delete()
    return HttpResponse("ok", content_type='application/json')

@api_view(['POST'])
def setDiagnostico(request):
    '''Funcion para modificar diagnosticos de un paciente.'''
    #mismo proceso de leer df
    data=request.data
    diagnostico1Cod=data["principal"]
    diagnostico2=data["secundarios"]
    idPaciente=data["id"]
    dias_estada=data["dias"]
    diag2_final = ""

    #lee los archivos cie10 y norma.
    path = os.path.dirname(os.path.realpath(__file__))
    archivo = path+'\CIE10-GRD.xlsm'
    cie10 = pd.read_excel(archivo, sheet_name='CIE10 MOD')
    norma = pd.read_excel(archivo, sheet_name='NORMA')

    # Se tranforma a numérico entero el IR GRD ya que lo toma con un .0 al final
    norma["IR-GRD CÓDIGO v2.3"] = pd.to_numeric(norma["IR-GRD CÓDIGO v2.3"], downcast='integer')
    #encuentra diagnosticos secundarios
    nombres_diags2 = []
    diagnostico2Cod=diagnostico2
    diagnostico2Json=[]
    if str(diagnostico2) == 'nan':
        diagnostico2 = []
        diagnostico2Cod=None
    else:
        diagnostico2 = str(diagnostico2).split(',')
        for diag in diagnostico2:
            condicion_diag2 = cie10.loc[:, 'CODIGO'] == diag
            diagnostico2_pd = cie10.loc[condicion_diag2]

            grd_diagnostico2 = diagnostico2_pd['GRD'].to_frame(name='GRD')
            sev_diagnostico2 = diagnostico2_pd['SEV'].to_frame(name = 'SEV')
            nombre_diagnostico2 = diagnostico2_pd['DIAGNOSTICO'].to_frame(name='DIAGNOSTICO')
            aux={}
            if grd_diagnostico2.size != 0:
                diagnostico_dos = nombre_diagnostico2['DIAGNOSTICO'].values[0]
                grd = str(grd_diagnostico2['GRD'].values[0])
                sev = str(sev_diagnostico2['SEV'].values[0])
                aux['codigo']=str(diag)
                aux['nombre']=str(diagnostico_dos)
                
                diagnostico2Json.append(aux)
            else:
                condicion = cie10.loc[:, 'CODIGO'] == diag+'.0'
                diagnostico2_pd = cie10.loc[condicion]
                grd_diagnostico2 = diagnostico2_pd['GRD'].to_frame(name='GRD')
                sev_diagnostico2 = diagnostico2_pd['SEV'].to_frame(name = 'SEV')
                nombre_diagnostico2 = diagnostico2_pd['DIAGNOSTICO'].to_frame(name='DIAGNOSTICO')
                diagnostico_dos = nombre_diagnostico2['DIAGNOSTICO'].values[0]
                grd = str(grd_diagnostico2['GRD'].values[0])
                sev = str(sev_diagnostico2['SEV'].values[0])
                aux['codigo']=str(diag)
                aux['nombre']=str(nombre_diagnostico2)
                
                diagnostico2Json.append(aux)


            nombres_diags2.append(diagnostico_dos)
            diag2_final=""
            for i in range(len(nombres_diags2)-1):
                diag2_final = diag2_final + nombres_diags2[i] +", "
            diag2_final = diag2_final + nombres_diags2[len(nombres_diags2)-1]  
    #encuentra diagnostico principal  
    if str(diagnostico1Cod) != 'nan':
        # Aquí busca el código del diagnóstico en el CIE10
        condicion = cie10.loc[:, 'CODIGO'] == diagnostico1Cod
        diagnostico1_pd = cie10.loc[condicion]

        grd_diagnostico1 = diagnostico1_pd['GRD'].to_frame(name='GRD')
        sev_diagnostico1 = diagnostico1_pd['SEV'].to_frame(name = 'SEV')
        nombre_diagnostico1 = diagnostico1_pd['DIAGNOSTICO'].to_frame(name='DIAGNOSTICO')
        if grd_diagnostico1.size != 0:
            diagnostico_uno = nombre_diagnostico1['DIAGNOSTICO'].values[0]
            grd = str(grd_diagnostico1['GRD'].values[0])
            sev = str(sev_diagnostico1['SEV'].values[0])
            
        else:
            condicion = cie10.loc[:, 'CODIGO'] == diagnostico1Cod+'.0'
            diagnostico1_pd = cie10.loc[condicion]
            grd_diagnostico1 = diagnostico1_pd['GRD'].to_frame(name='GRD')
            sev_diagnostico1 = diagnostico1_pd['SEV'].to_frame(name = 'SEV')
            nombre_diagnostico1 = diagnostico1_pd['DIAGNOSTICO'].to_frame(name='DIAGNOSTICO')
            diagnostico_uno = nombre_diagnostico1['DIAGNOSTICO'].values[0]
            grd = str(grd_diagnostico1['GRD'].values[0])
            sev = str(sev_diagnostico1['SEV'].values[0])
    else:
        diagnostico_uno = ""
        grd = ""
        sev = ""
        diagnostico1Cod=None
    # El GRD y SEV quedan como decimal por lo que se quita 
    # lo que está después del '.'
    if '.' in grd:
        i = grd.find('.')
        grd = grd[0:i]
    if '.' in sev:
        i2 = sev.find('.')
        sev = sev[0:i2]
    # Si la severidad es 0 o N entonces se añade un 1 al GRD
    if sev == '0' or sev == 'N':
        codigo_norma = str(grd)+'1'
    # Si la severidad es CC entonces se añade un 2 al GRD
    elif sev == 'CC':
        codigo_norma = str(grd)+'2'
    # Si la severidad es MCC entonces se añade un 3 al GRD
    elif sev == 'MCC':
        codigo_norma = str(grd)+'3'
    # Sino se añade un 1 al GRD
    else:
        codigo_norma = str(grd)+'1'

    # Ahora se busca el código IR-GRD en la norma
    norma["IR-GRD CÓDIGO v2.3"]=norma["IR-GRD CÓDIGO v2.3"].apply(str)
    condicion2 = norma.loc[:,'IR-GRD CÓDIGO v2.3'] == codigo_norma
    fila_norma = norma.loc[condicion2]
    pc_corte = 0
    peso_grd = 0
    em_norma = 0
    #si tiene norma
    if fila_norma.size == 0:
        print("No tiene NORMA")
    else:
        print("TIENE NORMA ")
        pc_corte = fila_norma['PC superior'].values[0]
        peso_grd = fila_norma['Peso GRD'].values[0]
        em_norma = fila_norma['EM \n(inlier)'].values[0]
    #actualiza los datos del paciente
    criterio = None
    if pc_corte!=0:
        criterio=float(dias_estada)/float(pc_corte)
    fecha = datetime.now()
    paciente=Resumen.objects.get(id=idPaciente)
    flag=True
    try:
        h = Historico.objects.get(fecha=paciente.updated_at, rut=paciente.rut, nombrePaciente=paciente.nombrePaciente,
        diagnostico1Cod=paciente.diagnostico1Cod, diagnostico2Cod=paciente.diagnostico2Cod,ir_grd=paciente.ir_grd,pesoGRD=paciente.pesoGRD,
        nombreServicio=paciente.nombreServicio,cama=paciente.cama, estancia=paciente.estancia)
    except Historico.DoesNotExist:
        h=None            
        print("no encuentra historico")
        flag=False
    #se guardan los nuevos diagnoticos y grupo grd
    paciente.pcSuperior=pc_corte
    paciente.pesoGRD=peso_grd
    paciente.emNorma=em_norma
    paciente.ir_grd=grd
    paciente.diagnostico1Cod=diagnostico1Cod
    paciente.diagnostico1=diagnostico_uno
    paciente.diagnostico2Cod=diagnostico2Cod
    paciente.diagnostico2=diag2_final
    paciente.diagnostico2Json=diagnostico2Json
    paciente.estancia=dias_estada
    paciente.criterio=criterio
    paciente.flag_diag=True
    paciente.updated_at=fecha
    paciente.save()
    #se guarda en el historico correspondiente.
    if flag:
        h.pcSuperior=pc_corte
        h.pesoGRD=peso_grd
        h.emNorma=em_norma
        h.ir_grd=grd
        h.diagnostico1Cod=diagnostico1Cod
        h.diagnostico1=diagnostico_uno
        h.diagnostico2Cod=diagnostico2Cod
        h.diagnostico2=diag2_final
        h.estancia=dias_estada
        h.criterio=criterio
        h.flag_diag=True
        h.save()
    return HttpResponse(data, content_type='application/json')
   


def estiloExcel(nombre,flag):
    '''Funcion para cambiar el estilo de un archivo excel.'''
    informe = openpyxl.load_workbook(nombre)
    sheet = informe.active 

    def set_width_to(sheet, start, stop, width):
        for col in get_column_interval(start, stop):
            sheet.column_dimensions[col].width = width
    
    # establece el tamaño de las celdas y los colores para que no se sobrepongan
    set_width_to(sheet, "A", "A", width=5)
    set_width_to(sheet, "B", "B", width=14)
    set_width_to(sheet, "C", "C", width=52)
    set_width_to(sheet, "D", "D", width=10)
    set_width_to(sheet, "E", "E", width=9)
    set_width_to(sheet, "F", "F", width=84)
    set_width_to(sheet, "G", "G", width=18)
    set_width_to(sheet, "H", "H", width=84)
    set_width_to(sheet, "I", "I", width=27)
    set_width_to(sheet, "J", "M", width=10)
    set_width_to(sheet, "N", "N", width=40)
    set_width_to(sheet, "O", "R", width=11)
    if flag:
        set_width_to(sheet, "S", "S", width=11)
        set_width_to(sheet, "T", "T", width=50)
        encabezados=["A1","B1","C1","D1","E1","F1","G1","H1","I1","J1","K1","L1","M1","N1","O1","P1","Q1","R1","S1","T1"]
    else:
        set_width_to(sheet, "S", "S", width=50)
        encabezados=["A1","B1","C1","D1","E1","F1","G1","H1","I1","J1","K1","L1","M1","N1","O1","P1","Q1","R1","S1"]
    #estilo para encabezados
    for encabezado in encabezados:
        celda = sheet[encabezado]
        celda.fill =  PatternFill("solid", fgColor="D9D9D9")
    informe.save(nombre)


def estiloExcelMensual(nombre,flag):
    '''Funcion para cambiar el estilo de un archivo excel de reporte mensual.'''
    informe = openpyxl.load_workbook(nombre)

    sheet = informe.active 

    def set_width_to(sheet, start, stop, width):
        for col in get_column_interval(start, stop):
            sheet.column_dimensions[col].width = width

    # establece el tamaño de las celdas y los colores para que no se sobrepongan
    set_width_to(sheet, "A", "A", width=5)
    set_width_to(sheet, "B", "B", width=40)
    set_width_to(sheet, "C", "C", width=10)
    set_width_to(sheet, "D", "D", width=15)
    set_width_to(sheet, "E", "N", width=22)
    encabezados=["A1","B1","C1","D1","E1","F1","G1","H1","I1","J1","K1","L1","M1","N1"]
    #estilo para encabezados
    for encabezado in encabezados:
        celda = sheet[encabezado]
        celda.fill =  PatternFill("solid", fgColor="D9D9D9")
    informe.save(nombre)



@api_view(['GET'])
def linkDescarga(request):
    '''Retorna una ruta de descarga para el archivo de resumen de pacientes'''
    nombreArchivo='Gestion_de_Pacientes.xlsx'
    nombreArchivoR='\Gestion_de_Pacientes.xlsx'
    ruta=os.path.dirname(os.path.abspath(__file__)) + nombreArchivoR
    with open(ruta,'rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s"' % nombreArchivo
    return response

@api_view(['GET'])
def linkDescargaH(request):
    '''Retorna una ruta de descarga para el archivo de historico de pacientes'''
    nombreArchivo='Historico_de_Pacientes.xlsx'
    nombreArchivoR='\Historico_de_Pacientes.xlsx'
    ruta=os.path.dirname(os.path.abspath(__file__)) + nombreArchivoR
    with open(ruta,'rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s"' % nombreArchivo
    return response


@api_view(['GET'])
def linkDescargaM(request):
    '''Retorna una ruta de descarga para el archivo de reporte mensual'''
    nombreArchivo='Reporte_mensual.xlsx'
    nombreArchivoR='\Reporte_mensual.xlsx'
    ruta=os.path.dirname(os.path.abspath(__file__)) + nombreArchivoR
    with open(ruta,'rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s"' % nombreArchivo
    return response

    
@api_view(['GET'])
def usuariosG(request):
    '''Get de los usuarios de la plataforma.'''
    users=User.objects.all()
    lista=[]
    for user in users:
        aux={}
        aux['id']=user.id
        aux["nombre_servicio"]=user.servicio.nombre
        aux["nombre_rol"]=user.rol.nombre
        aux["nombre"]= user.nombre
        aux["apellido"]= user.apellido
        aux["nickname"]= user.nickname
        aux["rut"]= user.rut
        lista.append(aux)
    return JsonResponse(lista, safe=False, json_dumps_params={'ensure_ascii':False})

@api_view(['POST'])   
def mensual_to_excel(request):
    '''Funcion genera un excel para el reporte mensual de entrada.'''    
    resumenJSON=request.data
    fecha=[]
    nombreServicio=[] 
    servicio=[] 
    em=[]
    emaf=[]
    iema=[]
    peso=[]
    iemaInliersMenor=[]
    iemaInliersMayor=[]
    outliers=[]
    pInt=[]
    pExt=[]
    condP=[]
    #para cada servicio en el resumen mensual
    for ser in resumenJSON:
        fecha.append(ser["fecha"])
        nombreServicio.append(ser['servicioNombre']) 
        servicio.append(ser['servicioId']) 
        em.append(ser["em"])
        emaf.append(ser["emaf"])
        iema.append(ser["iema"])
        peso.append(ser["peso"])
        iemaInliersMenor.append(ser["iemaInliersMenor"])
        iemaInliersMayor.append(ser["iemaInliersMayor"])
        outliers.append(ser["outliers"])
        pInt.append(ser["pInt"])
        pExt.append(ser["pExt"])
        condP.append(ser["condP"])
    #se crea un dataframe para el reporte mensual
    resumen= pd.DataFrame()
    resumen=resumen.assign(nombreServicio=nombreServicio,servicio=servicio,fecha=fecha,em=em,emaf=emaf,iema=iema,peso=peso,
    iemaInliersMenor=iemaInliersMenor,iemaInliersMayor=iemaInliersMayor,outliers=outliers,pInt=pInt,pExt=pExt,condP=condP)
    #crea el archivo y le da el estilo.
    nombreArchivo='gestionPacientes\Reporte_mensual.xlsx'
    resumen.to_excel(nombreArchivo, sheet_name='Resumen Mensual')
    estiloExcelMensual(nombreArchivo, True)
    resp={}
    resp["msg"]="Creado correctamente"
    return JsonResponse(resp, safe=False, json_dumps_params={'ensure_ascii':False})


@api_view(['POST'])   
def historico_to_excel(request):
    '''Funcion genera un excel para el reporte historico de entrada.'''    
    resumenJSON=request.data
    cama=[]
    rut=[]
    nombrePaciente=[] 
    estancia=[] 
    criterio=[] 
    diagnostico1=[] 
    diagnostico2 =[]
    diagnostico1Cod=[] 
    diagnostico2Cod =[]
    ir_grd=[] 
    emNorma=[] 
    pcSuperior=[] 
    pesoGRD=[] 
    nombreServicio=[] 
    servicio=[] 
    flag_diag =[]
    flag_pend=[]
    pendientesJson=[]
    fecha=[]
    #para cada paciente en el resumen
    for paciente in resumenJSON:
        cama.append(paciente['cama']) 
        rut.append(paciente['rut']) 
        nombrePaciente.append(paciente['nombrePaciente']) 
        estancia.append(paciente['estancia']) 
        criterio.append(paciente['criterio']) 
        diagnostico1.append(paciente['diagnostico1'])
        diagnostico1Cod.append(paciente['diagnostico1Cod'])  
        diagnostico2.append(paciente['diagnostico2'])
        diagnostico2Cod.append(paciente['diagnostico2Cod'])
        ir_grd.append(paciente['ir_grd']) 
        emNorma.append(paciente['emNorma']) 
        pcSuperior.append(paciente['pcSuperior']) 
        pesoGRD.append(paciente['pesoGRD']) 
        nombreServicio.append(paciente['nombreServicio']) 
        servicio.append(paciente['servicio_id']) 
        flag_diag.append(paciente['flag_diag'])
        flag_pend.append(paciente["flag_pend"])
        pend=""
        for j in paciente["pendientesJson"]:
            pend=pend+j["nombre"] + " ("+j["causa"]+"), \n"
        pendientesJson.append(pend)
        fecha.append(paciente["fecha"])

    #se crea un dataframe para el resumen
    resumen= pd.DataFrame()
    resumen=resumen.assign(rut=rut, nombrePaciente=nombrePaciente, cama=cama, estancia=estancia,
    diagnostico1=diagnostico1, diagnostico1Cod=diagnostico1Cod, diagnostico2=diagnostico2, diagnostico2Cod=diagnostico2Cod, ir_grd=ir_grd, emNorma=emNorma, 
    pcSuperior=pcSuperior, pesoGRD=pesoGRD, nombreServicio=nombreServicio, servicio=servicio, criterio=criterio, flag_diag=flag_diag,
    flag_pend=flag_pend, fecha=fecha,pendientesJson=pendientesJson)
    #crea el archivo y le da el estilo.
    nombreArchivo='gestionPacientes\Historico_de_Pacientes.xlsx'
    resumen.to_excel(nombreArchivo, sheet_name='Resumen de pacientes')
    estiloExcel(nombreArchivo, True)

    resp={}
    resp["msg"]="Creado correctamente"
    return JsonResponse(resp, safe=False, json_dumps_params={'ensure_ascii':False})

@api_view(['POST'])   
def resumen_to_excel(request):
    '''Funcion genera un excel para el resumen de pacientes de entrada.'''
    resumenJSON=request.data
    cama=[]
    rut=[]
    nombrePaciente=[] 
    estancia=[] 
    criterio=[] 
    diagnostico1=[] 
    diagnostico2 =[]
    diagnostico1Cod=[] 
    diagnostico2Cod =[]
    ir_grd=[] 
    emNorma=[] 
    pcSuperior=[] 
    pesoGRD=[] 
    nombreServicio=[] 
    servicio=[] 
    flag_diag =[]
    flag_pend=[]
    pendientesJson=[]
    #para cada paciente en el resumen
    for paciente in resumenJSON:
        cama.append(paciente['cama']) 
        rut.append(paciente['rut']) 
        nombrePaciente.append(paciente['nombrePaciente']) 
        estancia.append(paciente['estancia']) 
        criterio.append(paciente['criterio']) 
        diagnostico1.append(paciente['diagnostico1'])
        diagnostico1Cod.append(paciente['diagnostico1Cod'])  
        diagnostico2.append(paciente['diagnostico2'])
        diagnostico2Cod.append(paciente['diagnostico2Cod'])
        ir_grd.append(paciente['ir_grd']) 
        emNorma.append(paciente['emNorma']) 
        pcSuperior.append(paciente['pcSuperior']) 
        pesoGRD.append(paciente['pesoGRD']) 
        nombreServicio.append(paciente['nombreServicio']) 
        servicio.append(paciente['servicio']) 
        flag_diag.append(paciente['flag_diag'])
        flag_pend.append(paciente["flag_pend"])
        pend=""
        for j in paciente["pendientesJson"]:
            pend=pend+j["nombre"] + " ("+j["causa"]+"), \n"
        pendientesJson.append(pend)

    #se crea un dataframe para el resumen
    resumen= pd.DataFrame()
    resumen=resumen.assign(rut=rut, nombrePaciente=nombrePaciente, cama=cama, estancia=estancia,
    diagnostico1=diagnostico1, diagnostico1Cod=diagnostico1Cod, diagnostico2=diagnostico2, diagnostico2Cod=diagnostico2Cod, ir_grd=ir_grd, emNorma=emNorma, 
    pcSuperior=pcSuperior, pesoGRD=pesoGRD, nombreServicio=nombreServicio, servicio=servicio, criterio=criterio, flag_diag=flag_diag,
    flag_pend=flag_pend,pendientesJson=pendientesJson)
    #crea el archivo y le da el estilo.
    nombreArchivo='gestionPacientes\Gestion_de_Pacientes.xlsx'
    resumen.to_excel(nombreArchivo, sheet_name='Resumen de pacientes')
    estiloExcel(nombreArchivo,False)
    resp={}
    resp["msg"]="Creado correctamente"
    return JsonResponse(resp, safe=False, json_dumps_params={'ensure_ascii':False})


class UsuarioViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    queryset = User.objects.all()
    

class ServicioViewSet(viewsets.ModelViewSet):
    serializer_class = ServicioSerializer
    queryset = Servicio.objects.all().order_by('id')

class RolViewSet(viewsets.ModelViewSet):
    serializer_class = RolSerializer
    queryset = Roles.objects.all()

class PendienteViewSet(viewsets.ModelViewSet):
    serializer_class = PendienteSerializer
    queryset = Pendientes.objects.all()

# --- INTENTO DE FILTRO ---
from django_filters.rest_framework import DjangoFilterBackend
class ResumenViewSet(viewsets.ModelViewSet):
    '''serializer_class = ResumenSerializer
    queryset = Resumen.objects.all()
    filter_backends = [filters.SearchFilter]
    search_fields=['=servicio__id']'''
    queryset = Resumen.objects.all().order_by('-updated_at')#.values()
    serializer_class = ResumenSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['servicio__id', 'rut']


class HistoricoViewSet(viewsets.ModelViewSet):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['fecha']



class HistoricoDatesViewSet(viewsets.ModelViewSet):
    queryset = Historico.objects.values('fecha', 'id')
    serializer_class = HistoricoDatesSerializer

class MensualDatesViewSet(viewsets.ModelViewSet):
    queryset = ReporteMensual.objects.values('fecha').distinct()
    serializer_class = HistoricoDatesSerializer


    

#----------------INTENTO DE LOGIN --------------------
# Create your views here.
@api_view(['POST'])
def comprobar(request):
    print(request.data)
    user=request.data
    data = Usuarios.objects.get(nickname=user['nickname'])
    if data.password==user['password']:
        #return HttpResponse(data, content_type='application/json')
        user=[{'entra': 'SI', 'rol': data.rol.nombre}]
        return HttpResponse(user, content_type='application/json')
    user=[{'entra': 'NO'}]
    return HttpResponse(user, content_type='application/json')


def get_tokens_for_user(user):
    # Generate Token Manually
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
}

class UserRegistrationView(APIView):
  renderer_classes = [UserRenderer]
  def post(self, request, format=None):
    serializer = UserRegistrationSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = serializer.save()
    token = get_tokens_for_user(user)
    return Response({'token':token, 'msg':'Registration Successful'}, status=status.HTTP_201_CREATED)

class UserLoginView(APIView):
  renderer_classes = [UserRenderer]
  def post(self, request, format=None):
    serializer = UserLoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    nickname = serializer.data.get('nickname')
    password = serializer.data.get('password')
    try:
        usuario=User.objects.get(nickname=nickname)
    except User.DoesNotExist:
        return Response({'errors':{'nickname':['Nickname o password inválida']}}, status=status.HTTP_404_NOT_FOUND)
    rol_id=usuario.rol_id
    servicio_id=usuario.servicio_id
    rol=Roles.objects.get(id=rol_id)
    rol=rol.nombre
    servicio=Servicio.objects.get(id=servicio_id)
    servicio=servicio.nombre
    user = authenticate(nickname=nickname, password=password)
    if user is not None:
      token = get_tokens_for_user(user)
      return Response({'token':token, 'msg':'Login Success', 'rol':rol,'servicio':servicio,'servicio_id':servicio_id, 'inicial':nickname}, status=status.HTTP_200_OK)
    else:
      return Response({'errors':{'nickname':['Nickname o password inválida']}}, status=status.HTTP_404_NOT_FOUND)

class UserProfileView(APIView):
  renderer_classes = [UserRenderer]
  permission_classes = [IsAuthenticated]
  def get(self, request, format=None):
    serializer = UserProfileSerializer(request.user)
    return Response(serializer.data, status=status.HTTP_200_OK)

class UserChangePasswordView(APIView):
  renderer_classes = [UserRenderer]
  permission_classes = [IsAuthenticated]
  def post(self, request, format=None):
    serializer = UserChangePasswordSerializer(data=request.data, context={'user':request.user})
    serializer.is_valid(raise_exception=True)
    return Response({'msg':'Password Changed Successfully'}, status=status.HTTP_200_OK)


class SendPasswordResetEmailView(APIView):
  renderer_classes = [UserRenderer]
  def post(self, request, format=None):
    serializer = SendPasswordResetEmailSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    return Response({'msg':'Password Reset link send. Please check your Email'}, status=status.HTTP_200_OK)


class UserPasswordResetView(APIView):
  renderer_classes = [UserRenderer]
  def post(self, request, uid, token, format=None):
    serializer = UserPasswordResetSerializer(data=request.data, context={'uid':uid, 'token':token})
    serializer.is_valid(raise_exception=True)
    return Response({'msg':'Password Reset Successfully'}, status=status.HTTP_200_OK)